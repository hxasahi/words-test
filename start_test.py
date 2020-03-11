# encoding:cp936
from openpyxl import load_workbook
wb = load_workbook('F:\unit2.xlsx')
ws = wb.active
error_id=2
right_num = 0
error_num = 0
print('')
print('               欢迎使用单词自测软件')
print('')
print('       1 继续           exit 退出')
print('')
xuanze = raw_input('[提示]==>请输入数字选择：')
print('')
def word() :
    ceshi = 0
    nums = ws.max_row
    for i in range(2,nums+1) :
        print('             {}'.format(ws.cell(row=i,column=2).value))
        print('')
        n = raw_input('[提示]==>请输入单词：')
        print('')
        #if n != ws.cell(row=i,coulum=1).value :
        if n == ws.cell(row=i,column=1).value :
            ceshi = 1
            global error_id
            error_id=error_id+1
            global right_num
            right_num = right_num + 1
            if i == nums :
                print('[提示]==>测试结束')
                print('')
                print('[结束]==>正确次数{}'.format(right_num))
                print('')
                print('[结束]==>错误次数{}'.format(error_num))
                print('')
                pass
        elif n == ':quit':
            break
        else :
            print('[提示]==>输入错误')
            print('')
            error_word(error_id)
            error_id = error_id + 1
def error_word(error_id) :
        print('             {}'.format(ws.cell(row=error_id,column=2).value))
        print('')
        global error_num
        error_num = error_num + 1
        n = raw_input('[错误]==>请再次输入单词：')
        print('')
        if n == ws.cell(row=error_id,column=1).value :
            pass
        elif n == ':quit' :
            pass
        else :
            error_word(error_id)
        
while 1 :
    if xuanze == '1' or xuanze == '':
        word()
        
    elif xuanze == 'exit' :
        break
    print('       1 继续           exit 退出')
    print('')
    xuanze = raw_input('[提示]==>请输入数字选择：')
    print('')
